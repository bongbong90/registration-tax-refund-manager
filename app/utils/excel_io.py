"""거래처 마스터 엑셀 가져오기/내보내기."""
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.client_service import (
    list_clients,
    create_client,
    update_client,
    find_matching_clients,
)

EXCEL_HEADERS = [
    "거래처명",
    "법인등록번호",
    "사업자등록번호",
    "대표자",
    "주소",
    "이메일",
    "담당자명",
    "담당자전화",
    "비고",
]

FIELD_MAP = {
    "거래처명": "client_name",
    "법인등록번호": "corporation_no",
    "사업자등록번호": "business_no",
    "대표자": "representative_name",
    "주소": "address",
    "이메일": "email",
    "담당자명": "manager_name",
    "담당자전화": "manager_phone",
    "비고": "memo",
}


def export_to_excel(file_path: Path) -> int:
    """전체 거래처를 엑셀로 내보내기. 반환: 내보낸 건수."""
    clients = list_clients()

    wb = Workbook()
    ws = wb.active
    ws.title = "거래처"
    ws.append(EXCEL_HEADERS)

    for c in clients:
        ws.append(
            [
                c.client_name,
                c.corporation_no,
                c.business_no,
                c.representative_name,
                c.address,
                c.email,
                c.manager_name,
                c.manager_phone,
                c.memo,
            ]
        )

    # 컬럼 너비 자동 조정
    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(15, len(header) * 2)

    wb.save(file_path)
    return len(clients)


def import_from_excel(file_path: Path) -> dict:
    """
    엑셀에서 거래처 가져오기.
    동일 client_name 있으면 업데이트, 없으면 신규 생성.

    반환: {"created": N, "updated": N, "skipped": N, "errors": [...]}.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # 헤더 검증
    header_row = [cell.value for cell in ws[1]]
    if header_row[: len(EXCEL_HEADERS)] != EXCEL_HEADERS:
        raise ValueError(f"엑셀 헤더가 예상과 다릅니다. 첫 행에 {EXCEL_HEADERS} 가 있어야 합니다.")

    result = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:  # 거래처명 비어있으면 스킵
            result["skipped"] += 1
            continue

        try:
            data = {}
            for i, header in enumerate(EXCEL_HEADERS):
                if i < len(row):
                    value = row[i]
                    if value is not None:
                        data[FIELD_MAP[header]] = str(value).strip()

            client_name = data.get("client_name")
            if not client_name:
                result["skipped"] += 1
                continue

            # 동일 거래처명 있는지 확인
            existing = find_matching_clients(client_name)
            exact_match = next((c for c in existing if c.client_name == client_name), None)

            if exact_match:
                # 업데이트
                update_fields = {k: v for k, v in data.items() if k != "client_name"}
                if update_fields:
                    update_client(exact_match.id, **update_fields)
                    result["updated"] += 1
                else:
                    result["skipped"] += 1
            else:
                # 신규 생성
                create_client(**data)
                result["created"] += 1

        except Exception as e:
            result["errors"].append(f"{row_idx}행: {e}")

    return result
